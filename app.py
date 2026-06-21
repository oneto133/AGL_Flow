import csv
import ctypes
import ctypes.wintypes
import ipaddress
import json
import os
import re
import sys
import socket
import unicodedata
import threading
from datetime import datetime
from pathlib import Path

from fastapi import BackgroundTasks, FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from trello import ler_base_de_dados, enviar_trello
from dotenv import load_dotenv

load_dotenv()

TRELLO_KEY = os.getenv("TRELLO_API_KEY")
TRELLO_TOKEN = os.getenv("TRELLO_TOKEN")

MAPEAMENTO = {
    "celula_1": os.getenv("TRELLO_CELULA1_DZ"),
    "celula_2": os.getenv("TRELLO_CELULA2_DZ")
}

class DadosCartao(BaseModel):
    codigo: str
    quantidade: str
    op: str
    linhaCelula: str

try:
    from config_store import APP_HOME, CONFIG_PATH, DEFAULT_CONFIG, load_config, resolve_path, save_config
except ImportError:
    APP_HOME = Path(__file__).resolve().parent
    CONFIG_DIR = APP_HOME / "config"
    CONFIG_PATH = CONFIG_DIR / "app_config.json"
    DEFAULT_CONFIG = {
        "data_root": str(APP_HOME),
        "reposicao_csv": "csv/Reposicao e Diversos.csv",
        "labels_dir": "etiquetas",
        "report_dir": "relatorio",
        "base_file": "relatorio/base.xlsx",
        "printer_name": "",
        "active_label_profile": "40x25",
        "label_profiles": {
            "100x80": {
                "name": "100 mm x 80 mm",
                "width_mm": 100,
                "height_mm": 80,
                "columns": 1,
                "gap_mm": 3,
                "left_x_mm": 3,
                "left_y_mm": 3,
            "right_x_mm": 53,
            "right_y_mm": 3,
            "printer_name": "",
        },
            "50x25": {
                "name": "50 mm x 25 mm",
                "width_mm": 50,
                "height_mm": 25,
                "columns": 2,
                "gap_mm": 3,
                "left_x_mm": 3,
                "left_y_mm": 2,
            "right_x_mm": 56,
            "right_y_mm": 2,
            "printer_name": "",
        },
            "40x25": {
                "name": "40 mm x 25 mm",
                "width_mm": 40,
                "height_mm": 25,
                "columns": 2,
                "gap_mm": 3,
                "left_x_mm": 3,
                "left_y_mm": 2,
            "right_x_mm": 46,
            "right_y_mm": 2,
            "printer_name": "",
        },
        },

        # 40mm largura + 3mm gap
        "two_column_offset_dots": 344,

        # 40mm em 203 DPI
        "label_width_dots": 320,

        # 25mm em 203 DPI
        "label_height_dots": 200,

        "source_section_prefix": "reposicaoediversos",
    }

    def load_config():
        if not CONFIG_PATH.is_file():
            return DEFAULT_CONFIG.copy()
        try:
            import json

            loaded = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                return {**DEFAULT_CONFIG, **loaded}
        except (OSError, ValueError, json.JSONDecodeError):
            pass
        return DEFAULT_CONFIG.copy()

    def save_config(config):
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        import json

        CONFIG_PATH.write_text(
            json.dumps(config, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

    def resolve_path(value, base_dir=None):
        if value in (None, ""):
            return None
        path = Path(str(value).strip())
        if path.is_absolute():
            return path
        return (base_dir or APP_HOME) / path


RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", APP_HOME))
RAW_LABEL_EXTENSIONS = {".zpl", ".prn", ".epl"}
SOURCE_CSV_CODE_KEYS = ("codigo", "cod", "codproduto", "sku", "etiqueta", "unnamed1")
SOURCE_CSV_SOLD_KEY = "unnamed11"
DESCRIPTION_MAX_LINES = 3
LABEL_DPI = int(os.getenv("ZEBRA_LABEL_DPI", "203"))
LABEL_STORAGE_KEYS = (
    "name",
    "width_mm",
    "height_mm",
    "columns",
    "gap_mm",
    "left_x_mm",
    "left_y_mm",
    "right_x_mm",
    "right_y_mm",
    "printer_name",
)

app = FastAPI(title="Impressao de Etiquetas")
style_dir = RESOURCE_DIR / "style"
images_dir = RESOURCE_DIR / "imagens"
if style_dir.is_dir():
    app.mount("/style", StaticFiles(directory=style_dir), name="style")
if images_dir.is_dir():
    app.mount("/imagens", StaticFiles(directory=images_dir), name="imagens")
templates = Jinja2Templates(directory=RESOURCE_DIR / "templates")

RUNTIME_CONFIG = load_config()
PRINT_STATS_PATH = APP_HOME / "config" / "print_stats.json"
PRINT_STATS_LOCK = threading.Lock()
SERVER_PORT = int(os.getenv("PORT", "8000"))
class PrintRequest(BaseModel):
    etiqueta: str = Field(min_length=1)
    quantidade: int = Field(ge=1, le=500)


class ConfigRequest(BaseModel):
    data_root: str | None = None
    reposicao_csv: str | None = None
    labels_dir: str | None = None
    report_dir: str | None = None
    base_file: str | None = None
    printer_name: str | None = None
    two_column_offset_dots: int | None = Field(default=None, ge=0, le=4000)
    label_width_dots: int | None = Field(default=None, ge=1, le=4000)
    label_height_dots: int | None = Field(default=None, ge=1, le=4000)
    source_section_prefix: str | None = None


class LabelProfileRequest(BaseModel):
    name: str | None = None
    width_mm: float = Field(gt=0, le=500)
    height_mm: float = Field(gt=0, le=500)
    columns: int = Field(ge=1, le=2)
    gap_mm: float = Field(ge=0, le=100)
    left_x_mm: float = Field(ge=-500, le=500)
    left_y_mm: float = Field(ge=-500, le=500)
    right_x_mm: float | None = Field(default=None, ge=-500, le=500)
    right_y_mm: float | None = Field(default=None, ge=-500, le=500)
    printer_name: str | None = None


class LabelConfigRequest(BaseModel):
    active_profile: str
    profiles: dict[str, LabelProfileRequest]


def get_config():
    return normalize_config({**DEFAULT_CONFIG, **RUNTIME_CONFIG})


def set_config(config):
    global RUNTIME_CONFIG
    RUNTIME_CONFIG = normalize_config({**DEFAULT_CONFIG, **config})
    save_config(RUNTIME_CONFIG)
    return get_config()


def mm_to_dots(value, dpi=LABEL_DPI):
    return int(round(float(value) * float(dpi) / 25.4))


def dots_to_mm(value, dpi=LABEL_DPI):
    return round(float(value) * 25.4 / float(dpi), 2)


def _default_label_profiles():
    return json.loads(json.dumps(DEFAULT_CONFIG["label_profiles"]))


def _coerce_float(value, default=0.0):
    try:
        if value in (None, ""):
            return float(default)
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _coerce_int(value, default=0):
    try:
        if value in (None, ""):
            return int(default)
        return int(round(float(value)))
    except (TypeError, ValueError):
        return int(default)


def normalize_label_profile(profile_id, profile):
    defaults = DEFAULT_CONFIG["label_profiles"].get(profile_id, {})
    merged = {**defaults, **(profile or {})}

    width_mm = max(_coerce_float(merged.get("width_mm"), defaults.get("width_mm", 0)), 0.1)
    height_mm = max(_coerce_float(merged.get("height_mm"), defaults.get("height_mm", 0)), 0.1)
    columns = 2 if _coerce_int(merged.get("columns"), defaults.get("columns", 1)) >= 2 else 1
    gap_mm = max(_coerce_float(merged.get("gap_mm"), defaults.get("gap_mm", 3)), 0.0)
    left_x_mm = _coerce_float(merged.get("left_x_mm"), defaults.get("left_x_mm", 0))
    left_y_mm = _coerce_float(merged.get("left_y_mm"), defaults.get("left_y_mm", 0))

    if columns > 1:
        right_x_default = left_x_mm + width_mm + gap_mm
        right_y_default = left_y_mm
        right_x_mm = _coerce_float(merged.get("right_x_mm"), right_x_default)
        right_y_mm = _coerce_float(merged.get("right_y_mm"), right_y_default)
    else:
        right_x_mm = _coerce_float(merged.get("right_x_mm"), left_x_mm)
        right_y_mm = _coerce_float(merged.get("right_y_mm"), left_y_mm)

    return {
        "id": profile_id,
        "name": str(merged.get("name") or defaults.get("name") or profile_id).strip(),
        "width_mm": round(width_mm, 2),
        "height_mm": round(height_mm, 2),
        "columns": columns,
        "gap_mm": round(gap_mm, 2),
        "left_x_mm": round(left_x_mm, 2),
        "left_y_mm": round(left_y_mm, 2),
        "right_x_mm": round(right_x_mm, 2),
        "right_y_mm": round(right_y_mm, 2),
        "printer_name": str(merged.get("printer_name") or "").strip(),
    }


def get_label_profiles(config=None):
    config = {**DEFAULT_CONFIG, **(config or get_config())}
    stored_profiles = config.get("label_profiles")
    profiles = _default_label_profiles()

    if isinstance(stored_profiles, dict):
        for profile_id, profile in stored_profiles.items():
            if isinstance(profile, dict):
                profiles[profile_id] = {**profiles.get(profile_id, {}), **profile}

    return {profile_id: normalize_label_profile(profile_id, profile) for profile_id, profile in profiles.items()}


def get_active_label_profile_id(config=None, profiles=None):
    config = {**DEFAULT_CONFIG, **(config or get_config())}
    profiles = profiles or get_label_profiles(config)
    active_id = str(config.get("active_label_profile") or "").strip()
    if active_id in profiles:
        return active_id
    return next(iter(profiles))


def get_label_layout(config=None):
    config = {**DEFAULT_CONFIG, **(config or get_config())}
    profiles = get_label_profiles(config)
    profile_id = get_active_label_profile_id(config, profiles)
    profile = profiles[profile_id]

    width_dots = mm_to_dots(profile["width_mm"])
    height_dots = mm_to_dots(profile["height_mm"])
    left_x_dots = mm_to_dots(profile["left_x_mm"])
    left_y_dots = mm_to_dots(profile["left_y_mm"])
    right_x_dots = mm_to_dots(profile["right_x_mm"])
    right_y_dots = mm_to_dots(profile["right_y_mm"])
    gap_dots = mm_to_dots(profile["gap_mm"])
    column_positions = [{"x": left_x_dots, "y": left_y_dots}]
    if profile["columns"] > 1:
        column_positions.append({"x": right_x_dots, "y": right_y_dots})
    min_x_dots = min(position["x"] for position in column_positions)
    max_x_dots = max(position["x"] + width_dots for position in column_positions)
    page_width_dots = max(width_dots, max_x_dots - min(0, min_x_dots) + 320)

    return {
        **profile,
        "dpi": LABEL_DPI,
        "width_dots": width_dots,
        "height_dots": height_dots,
        "page_width_dots": page_width_dots,
        "left_x_dots": left_x_dots,
        "left_y_dots": left_y_dots,
        "right_x_dots": right_x_dots,
        "right_y_dots": right_y_dots,
        "gap_dots": gap_dots,
        "two_column_offset_dots": right_x_dots - left_x_dots,
        "column_positions": column_positions,
    }


def get_effective_printer_name(config=None):
    config = normalize_config({**DEFAULT_CONFIG, **(config or get_config())})
    profiles = get_label_profiles(config)
    active_id = get_active_label_profile_id(config, profiles)
    profile_printer = str(profiles[active_id].get("printer_name") or "").strip()
    global_printer = str(config.get("printer_name") or "").strip()
    return profile_printer or global_printer


def list_available_printers():
    winspool = ctypes.WinDLL("winspool.drv")
    PRINTER_ENUM_LOCAL = 0x00000002
    PRINTER_ENUM_CONNECTIONS = 0x00000004

    class PRINTER_INFO_4(ctypes.Structure):
        _fields_ = [
            ("pPrinterName", ctypes.wintypes.LPWSTR),
            ("pServerName", ctypes.wintypes.LPWSTR),
            ("Attributes", ctypes.wintypes.DWORD),
        ]

    needed = ctypes.wintypes.DWORD(0)
    returned = ctypes.wintypes.DWORD(0)
    flags = PRINTER_ENUM_LOCAL | PRINTER_ENUM_CONNECTIONS

    winspool.EnumPrintersW(flags, None, 4, None, 0, ctypes.byref(needed), ctypes.byref(returned))
    if needed.value == 0:
        default_printer = ""
        try:
            default_printer = get_default_printer_name()
        except Exception:
            default_printer = ""
        return [default_printer] if default_printer else []

    buffer = ctypes.create_string_buffer(needed.value)
    if not winspool.EnumPrintersW(
        flags,
        None,
        4,
        buffer,
        needed.value,
        ctypes.byref(needed),
        ctypes.byref(returned),
    ):
        raise ctypes.WinError()

    printers = []
    array_type = PRINTER_INFO_4 * returned.value
    printer_array = ctypes.cast(buffer, ctypes.POINTER(array_type)).contents
    for printer in printer_array:
        name = (printer.pPrinterName or "").strip()
        if name and name not in printers:
            printers.append(name)

    try:
        default_printer = get_default_printer_name()
    except Exception:
        default_printer = ""
    if default_printer and default_printer not in printers:
        printers.insert(0, default_printer)

    return printers


def normalize_config(config):
    normalized = {**DEFAULT_CONFIG, **(config or {})}
    profiles = get_label_profiles(normalized)
    active_id = get_active_label_profile_id(normalized, profiles)
    layout = get_label_layout({**normalized, "label_profiles": profiles, "active_label_profile": active_id})
    normalized["label_profiles"] = {
        profile_id: {key: profile[key] for key in LABEL_STORAGE_KEYS}
        for profile_id, profile in profiles.items()
    }
    normalized["active_label_profile"] = active_id
    normalized["two_column_offset_dots"] = layout["two_column_offset_dots"]
    normalized["label_width_dots"] = layout["width_dots"]
    normalized["label_height_dots"] = layout["height_dots"]
    normalized["label_dpi"] = LABEL_DPI
    return normalized


RUNTIME_CONFIG = normalize_config(RUNTIME_CONFIG)
if not CONFIG_PATH.exists():
    save_config(RUNTIME_CONFIG)


def get_paths(config=None):
    config = normalize_config({**DEFAULT_CONFIG, **(config or get_config())})
    layout = get_label_layout(config)
    data_root = resolve_path(config.get("data_root") or DEFAULT_CONFIG["data_root"], APP_HOME) or APP_HOME
    active_profiles = get_label_profiles(config)
    active_profile_id = get_active_label_profile_id(config, active_profiles)
    return {
        "data_root": data_root,
        "reposicao_csv": resolve_path(config.get("reposicao_csv"), data_root) or data_root / "csv" / "Reposicao e Diversos.csv",
        "labels_dir": resolve_path(config.get("labels_dir"), data_root) or data_root / "etiquetas",
        "report_dir": resolve_path(config.get("report_dir"), data_root) or data_root / "relatorio",
        "base_file": resolve_path(config.get("base_file"), data_root) or data_root / "relatorio" / "base.xlsx",
        "printer_name": get_effective_printer_name(config),
        "global_printer_name": (config.get("printer_name") or "").strip(),
        "profile_printer_name": str(active_profiles[active_profile_id].get("printer_name") or "").strip(),
        "two_column_offset_dots": layout["two_column_offset_dots"],
        "label_width_dots": layout["width_dots"],
        "label_height_dots": layout["height_dots"],
        "label_page_width_dots": layout["page_width_dots"],
        "label_dpi": LABEL_DPI,
        "label_profile_id": layout["id"],
        "label_profile_name": layout["name"],
        "label_columns": layout["columns"],
        "label_column_positions_dots": layout["column_positions"],
        "label_left_x_dots": layout["left_x_dots"],
        "label_left_y_dots": layout["left_y_dots"],
        "label_right_x_dots": layout["right_x_dots"],
        "label_right_y_dots": layout["right_y_dots"],
        "source_section_prefix": (config.get("source_section_prefix") or DEFAULT_CONFIG["source_section_prefix"]).strip().lower(),
    }


def ensure_parent_dirs(paths):
    paths["report_dir"].mkdir(parents=True, exist_ok=True)
    paths["labels_dir"].mkdir(parents=True, exist_ok=True)
    paths["reposicao_csv"].parent.mkdir(parents=True, exist_ok=True)
    paths["base_file"].parent.mkdir(parents=True, exist_ok=True)


def get_reposicao_metadata(config=None):
    paths = get_paths(config)
    csv_path = paths["reposicao_csv"]
    exists = csv_path.is_file()
    updated_at = ""
    updated_at_iso = ""
    size_bytes = 0

    if exists:
      # CSV is kept external and updated daily, so we only read metadata here.
        stat = csv_path.stat()
        updated = datetime.fromtimestamp(stat.st_mtime)
        updated_at = updated.strftime("%d/%m/%Y %H:%M:%S")
        updated_at_iso = updated.isoformat(timespec="seconds")
        size_bytes = stat.st_size

    return {
        "path": str(csv_path),
        "exists": exists,
        "updated_at": updated_at,
        "updated_at_iso": updated_at_iso,
        "size_bytes": size_bytes,
    }


def _normalize_key(value):
    normalized = unicodedata.normalize("NFKD", value)
    plain = "".join(char for char in normalized if not unicodedata.combining(char)).lower()
    return re.sub(r"[^a-z0-9]", "", plain)


def value_to_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def only_digits(value):
    return re.sub(r"\D", "", value or "")


def strip_accents(value):
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(char for char in normalized if not unicodedata.combining(char))


def zpl_text(value):
    text = strip_accents(value).upper()
    return text.replace("\\", " ").replace("^", " ").replace("~", " ")


def numeric_quantity(value):
    if isinstance(value, (int, float)):
        return int(abs(value)) if value else 0

    text = value_to_text(value).replace(".", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return 0

    return int(abs(number)) if number else 0


def _normalize_print_stats(data):
    items = {}
    raw_items = data.get("items") if isinstance(data, dict) and isinstance(data.get("items"), dict) else data

    if not isinstance(raw_items, dict):
        raw_items = {}

    for code, value in raw_items.items():
        code_text = value_to_text(code)
        if not code_text:
            continue

        if isinstance(value, dict):
            count = numeric_quantity(value.get("count"))
            last_printed_at = value.get("last_printed_at") or ""
        else:
            count = numeric_quantity(value)
            last_printed_at = ""

        items[code_text] = {
            "count": count,
            "last_printed_at": value_to_text(last_printed_at),
        }

    return {"items": items}


def load_print_stats():
    if not PRINT_STATS_PATH.is_file():
        return {"items": {}}

    try:
        loaded = json.loads(PRINT_STATS_PATH.read_text(encoding="utf-8"))
        if isinstance(loaded, dict):
            return _normalize_print_stats(loaded)
    except (OSError, ValueError, json.JSONDecodeError):
        pass

    return {"items": {}}


def save_print_stats(stats):
    PRINT_STATS_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRINT_STATS_PATH.write_text(
        json.dumps(stats, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


PRINT_STATS = load_print_stats()


def get_printed_count(codigo):
    code_text = value_to_text(codigo)
    return numeric_quantity(PRINT_STATS.get("items", {}).get(code_text, {}).get("count", 0))


def record_printed_quantity(codigo, quantidade):
    global PRINT_STATS
    code_text = value_to_text(codigo)
    if not code_text:
        return

    with PRINT_STATS_LOCK:
        stats = _normalize_print_stats(PRINT_STATS)
        item = stats.setdefault("items", {}).get(code_text, {"count": 0, "last_printed_at": ""})
        item["count"] = numeric_quantity(item.get("count", 0)) + numeric_quantity(quantidade)
        item["last_printed_at"] = datetime.now().isoformat(timespec="seconds")
        stats["items"][code_text] = item

        PRINT_STATS = stats
        save_print_stats(PRINT_STATS)


def find_base_header(worksheet):
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [_normalize_key(value_to_text(value)) for value in row]
        has_code = "codigo" in normalized or "cod" in normalized
        has_barcode = bool({"codigobarrasean", "codigobarras", "ean", "barcode"} & set(normalized))
        if has_code and has_barcode:
            return row_index, normalized

    raise RuntimeError("Cabecalho da base nao encontrado.")


def find_header_index(headers, names):
    for index, header in enumerate(headers):
        if header in names:
            return index
    return None


def read_reposicao_rows(config=None):
    paths = get_paths(config)
    csv_path = paths["reposicao_csv"]
    if not csv_path.exists() or csv_path.stat().st_size == 0:
        return []

    with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as file:
        sample = file.read(4096)
        file.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except csv.Error:
            dialect = csv.excel

        try:
            has_header = csv.Sniffer().has_header(sample) if sample.strip() else False
        except csv.Error:
            has_header = False

        if has_header:
            reader = csv.DictReader(file, dialect=dialect)
            rows = [_row_from_source_dict(row, paths) for row in reader]
            return [row for row in rows if row]

        reader = csv.reader(file, dialect=dialect)
        rows = [_row_from_source_list(index, row, paths) for index, row in enumerate(reader, start=1)]
        return [row for row in rows if row]


def _row_from_source_dict(row, paths):
    clean_row = {str(key).strip(): value_to_text(value) for key, value in row.items() if key}
    normalized_row = {_normalize_key(key): value for key, value in clean_row.items()}
    codigo = next((normalized_row[key] for key in SOURCE_CSV_CODE_KEYS if normalized_row.get(key)), "")
    descricao = next((normalized_row[key] for key in ("descricao", "produto", "nome") if normalized_row.get(key)), "")
    secao = normalized_row.get("secao", "")
    quantidade_vendida = normalized_row.get(SOURCE_CSV_SOLD_KEY, "")

    if secao and not _normalize_key(secao).startswith(paths["source_section_prefix"]):
        return None

    if not quantidade_vendida:
        quantidade_vendida = next(
            (
                normalized_row[key]
                for key in ("quantidadevendida", "quantidade", "qtd", "qtde")
                if normalized_row.get(key)
            ),
            "",
        )

    if not codigo:
        codigo = _find_matching_code(clean_row.values(), paths["labels_dir"]) or next((value for value in clean_row.values() if value), "")

    return {
        "codigo": codigo,
        "nome": descricao or codigo,
        "secao": secao,
        "quantidade_vendida": numeric_quantity(quantidade_vendida),
        "dados": clean_row,
    }


def _row_from_source_list(index, row, paths):
    values = [value_to_text(value) for value in row]
    codigo = value_to_text(row[1] if len(row) > 1 else "")
    descricao = value_to_text(row[2] if len(row) > 2 else "")
    secao = value_to_text(row[3] if len(row) > 3 else "")
    quantidade_vendida = value_to_text(row[11] if len(row) > 11 else "")

    if secao and not _normalize_key(secao).startswith(paths["source_section_prefix"]):
        return None

    if not codigo:
        non_empty = [value for value in values if value]
        codigo = _find_matching_code(non_empty, paths["labels_dir"]) or (non_empty[0] if non_empty else f"Linha {index}")

    return {
        "codigo": codigo,
        "nome": descricao or codigo,
        "secao": secao,
        "quantidade_vendida": numeric_quantity(quantidade_vendida),
        "dados": {"linha": index, "valores": values},
    }


def _find_matching_code(values, labels_dir):
    label_codes = {path.stem for path in labels_dir.glob("*") if path.is_file()}
    for value in values:
        cleaned = Path(str(value).strip()).stem
        if cleaned in label_codes:
            return cleaned
    return ""


def find_label_file(nome, labels_dir=None):
    paths = get_paths()
    labels_dir = labels_dir or paths["labels_dir"]
    safe_name = Path(nome).stem
    matches = [path for path in labels_dir.glob("*") if path.is_file() and path.stem == safe_name]
    raw_matches = [path for path in matches if path.suffix.lower() in RAW_LABEL_EXTENSIONS]
    return (raw_matches or matches or [None])[0]


def is_automatic_label(label_path):
    return label_path.suffix.lower() in RAW_LABEL_EXTENSIONS


def get_base_file(paths=None):
    paths = paths or get_paths()
    base_file = paths["base_file"]
    if base_file.is_file() and not base_file.name.startswith("~$"):
        return base_file

    report_dir = paths["report_dir"]
    if not report_dir.exists():
        return None

    files = [
        path
        for path in report_dir.glob("*.xls*")
        if path.is_file() and not path.name.startswith("~$") and "base" in path.stem.lower()
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def read_base_items(base_path):
    workbook = load_workbook(base_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_row_index, headers = find_base_header(worksheet)

        codigo_col = find_header_index(headers, {"codigo", "cod"})
        descricao_col = find_header_index(headers, {"descricao"})
        barcode_col = find_header_index(headers, {"codigobarrasean", "codigobarras", "ean", "barcode"})

        if codigo_col is None or barcode_col is None:
            raise RuntimeError("Base precisa ter as colunas Codigo e Codigo Barras EAN.")

        items = {}
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            codigo = value_to_text(row[codigo_col] if codigo_col < len(row) else "")
            if not codigo:
                continue

            items[codigo] = {
                "descricao": value_to_text(row[descricao_col] if descricao_col is not None and descricao_col < len(row) else ""),
                "codigo_barras": only_digits(value_to_text(row[barcode_col] if barcode_col < len(row) else "")),
            }

        return items
    finally:
        workbook.close()


def get_base_item(codigo):
    paths = get_paths()
    base_path = get_base_file(paths)
    if base_path is None:
        return None

    base_item = read_base_items(base_path).get(value_to_text(codigo))
    if not base_item:
        return None

    return {
        "codigo": value_to_text(codigo),
        "nome": base_item.get("descricao", ""),
        "secao": "",
        "codigo_barras": base_item.get("codigo_barras", ""),
        "quantidade_vendida": "",
        "dados": {},
        "etiqueta": "dinamica.zpl" if base_item.get("codigo_barras") else "",
        "disponivel": bool(base_item.get("codigo_barras")),
        "automatico": bool(base_item.get("codigo_barras")),
    }


def get_items():
    paths = get_paths()
    items = []
    seen = set()
    base_path = get_base_file(paths)
    base_items = read_base_items(base_path) if base_path is not None else {}

    for row in read_reposicao_rows(paths):
        codigo = row["codigo"].strip()
        if not codigo or codigo in seen:
            continue

        base_item = base_items.get(codigo, {})
        codigo_barras = base_item.get("codigo_barras", "")
        nome = base_item.get("descricao") or row.get("nome", codigo)
        label_file = find_label_file(codigo, paths["labels_dir"])
        items.append(
            {
                "codigo": codigo,
                "nome": nome,
                "secao": row.get("secao", ""),
                "codigo_barras": codigo_barras,
                "quantidade_vendida": row.get("quantidade_vendida", ""),
                "quantidade_impresso": get_printed_count(codigo),
                "dados": row["dados"],
                "etiqueta": "dinamica.zpl" if codigo_barras else label_file.name if label_file else "",
                "disponivel": bool(codigo_barras) or label_file is not None,
                "automatico": bool(codigo_barras) or (is_automatic_label(label_file) if label_file else False),
            }
        )
        seen.add(codigo)

    return items


def get_lan_ip():
    candidates = []

    try:
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None, socket.AF_INET, socket.SOCK_STREAM):
            candidate = info[4][0]
            if candidate and candidate not in candidates:
                candidates.append(candidate)
    except OSError:
        pass

    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            ip_address = sock.getsockname()[0]
            if ip_address and ip_address not in candidates:
                candidates.append(ip_address)
    except OSError:
        pass

    for candidate in candidates:
        try:
            ip_obj = ipaddress.ip_address(candidate)
        except ValueError:
            continue

        if ip_obj.version == 4 and not (ip_obj.is_loopback or ip_obj.is_link_local or ip_obj.is_unspecified):
            if ip_obj.is_private:
                return candidate

    for candidate in candidates:
        try:
            ip_obj = ipaddress.ip_address(candidate)
        except ValueError:
            continue

        if ip_obj.version == 4 and not (ip_obj.is_loopback or ip_obj.is_link_local or ip_obj.is_unspecified):
            return candidate

    return "127.0.0.1"


def get_access_url():
    return f"http://{get_lan_ip()}:{SERVER_PORT}"


def _port_is_available(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            sock.bind(("0.0.0.0", port))
        except OSError:
            return False
    return True


def find_available_port(start=8000, end=8002):
    for port in range(start, end + 1):
        if _port_is_available(port):
            return port
    raise OSError(f"Nenhuma porta disponivel entre {start} e {end}.")


def queue_print_item(background_tasks, item, quantidade):
    paths = get_paths()
    if item.get("codigo_barras"):
        printer_name = paths["printer_name"] or get_default_printer_name()
        background_tasks.add_task(print_dynamic_item, item, quantidade, printer_name)
        return {
            "message": "Impressao dinamica enviada.",
            "etiqueta": item["codigo"],
            "arquivo": "dinamica.zpl",
            "quantidade": quantidade,
            "impressora": printer_name,
        }

    label_file = find_label_file(item["codigo"], paths["labels_dir"])
    if label_file is None:
        raise HTTPException(
            status_code=404,
            detail="Codigo de barras ou arquivo de etiqueta com o mesmo nome nao foi encontrado.",
        )

    if not is_automatic_label(label_file):
        raise HTTPException(
            status_code=400,
            detail=(
                "Impressao automatica sem abrir ZebraDesigner precisa de arquivo .zpl, .prn ou .epl "
                f"com o mesmo codigo. Arquivo atual: {label_file.name}."
            ),
        )

    printer_name = paths["printer_name"] or get_default_printer_name()
    background_tasks.add_task(print_label_file, label_file, quantidade, printer_name)
    return {
        "message": "Impressao enviada pelo ZebraDesigner.",
        "etiqueta": item["codigo"],
        "arquivo": label_file.name,
        "quantidade": quantidade,
        "impressora": printer_name,
    }


def print_sold_items_batch():
    printer_name = get_paths()["printer_name"] or get_default_printer_name()
    sold_items = [item for item in get_items() if numeric_quantity(item.get("quantidade_vendida")) > 0 and item.get("automatico")]
    results = []
    for item in sold_items:
        quantidade = numeric_quantity(item.get("quantidade_vendida"))
        try:
            if item.get("codigo_barras"):
                print_dynamic_item(item, quantidade, printer_name)
                results.append({"codigo": item["codigo"], "quantidade": quantidade, "arquivo": "dinamica.zpl"})
            else:
                label_file = find_label_file(item["codigo"], get_paths()["labels_dir"])
                if label_file and is_automatic_label(label_file):
                    print_label_file(label_file, quantidade, printer_name)
                    results.append({"codigo": item["codigo"], "quantidade": quantidade, "arquivo": label_file.name})
        except Exception as error:
            results.append({"codigo": item["codigo"], "erro": str(error)})

    return {
        "total_enviados": sum(1 for result in results if "erro" not in result),
        "total_itens": len(sold_items),
        "itens": results,
    }


def print_label_file(label_path, quantidade, printer_name=None):
    print_raw_label(label_path, quantidade, printer_name)


def print_dynamic_item(item, quantidade, printer_name=None):
    config = get_paths()
    printer_name = printer_name or config["printer_name"] or get_default_printer_name()
    data = build_dynamic_zpl(item, quantidade, config)
    send_raw_to_printer(printer_name, data, f"{item['codigo']} x{quantidade}")
    record_printed_quantity(item["codigo"], quantidade)


def build_dynamic_zpl(item, quantidade, config):
    columns = max(int(config.get("label_columns", 1) or 1), 1)
    parts = [b"\xef\xbb\xbf"]

    if columns == 1:
        parts.append(build_dynamic_zpl_block(item, quantidade, columns=1, config=config))
    else:
        pair_rows = quantidade // 2
        single_rows = quantidade % 2

        if pair_rows:
            parts.append(build_dynamic_zpl_block(item, pair_rows, columns=2, config=config))

        if single_rows:
            parts.append(build_dynamic_zpl_block(item, single_rows, columns=1, config=config))

    return b"".join(parts)


def build_dynamic_zpl_block(item, rows, columns, config):
    layout_positions = config.get("label_column_positions_dots") or [{"x": 0, "y": 0}]
    positions = layout_positions[:columns]
    if len(positions) < columns:
        positions = positions + [{"x": 0, "y": 0}] * (columns - len(positions))

    commands = [
        "^XA",

        # modo térmico
        "^MMT",

        # largura total da mídia
        f"^PW{config.get('label_page_width_dots', 720)}",

        # altura da etiqueta
        f"^LL{config['label_height_dots']}",

        # deslocamento global horizontal
        "^LS0",

        # UTF-8
        "^CI28",
    ]

    for column in range(columns):
        position = positions[column]
        commands.extend(
            build_label_column(
                item,
                position,
                config,
            )
        )

    commands.extend([
        f"^PQ{rows},0,1,Y",
        "^XZ"
    ])

    return ("\r\n".join(commands) + "\r\n").encode("utf-8")


def _build_description_layout(value):
    description = re.sub(r"\s+", " ", zpl_text(value or "")).strip()
    length = len(description)
    estimated_lines = min(3, max(1, (length + 27) // 28))

    if estimated_lines >= 3:
        font_size = 18
        box_width = 232
    elif estimated_lines == 2:
        font_size = 20
        box_width = 242
    else:
        font_size = 22
        box_width = 250

    return description, font_size, box_width


def build_label_column(item, position, config):
    x_offset = int(position.get("x", 0))
    y_offset = int(position.get("y", 0))

    description, description_font, description_width = _build_description_layout(item.get("nome") or "")
    code = zpl_text(item.get("codigo") or "")
    barcode = only_digits(item.get("codigo_barras") or "")
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M")

    commands = [

        # =========================
        # DESCRIÇÃO
        # =========================
        f"""
            ^FO{x_offset + 22},{y_offset + 28}
            ^A0N,{description_font},{description_font}
            ^FB{description_width},3,1,L,0
            ^FD{description}^FS
            """.strip(),

        # =========================
        # CÓDIGO
        # =========================
        f"""
            ^FO{x_offset + 22},{y_offset + 96}
            ^A0N,22,22
            ^FD{code}^FS
            """.strip(),

        # =========================
        # DATA
        # =========================
        f"""
            ^FO{x_offset + 165},{y_offset + 96}
            ^A0N,16,16
            ^FD{current_date}^FS
            """.strip(),
                ]

    if barcode:

        commands.extend([

            # =========================
            # CÓDIGO DE BARRAS
            # =========================
            f"""
            ^BY2,2,52
            ^FT{x_offset + 58},{y_offset + 178}
            ^BEN,52,Y,N
            ^FD{barcode}^FS
            """.strip()

                    ])

    else:

        commands.append(

            f"""
                ^FO{x_offset + 22},{y_offset + 145}
                ^A0N,20,20
                ^FDEAN NAO ENCONTRADO^FS
                """.strip()

                        )

    return commands


def print_raw_label(label_path, quantidade, printer_name=None):
    config = get_paths()
    printer_name = printer_name or config["printer_name"] or get_default_printer_name()
    original_data = label_path.read_bytes()

    if not original_data:
        raise RuntimeError(f"Arquivo vazio: {label_path.name}")

    columns = max(int(config.get("label_columns", 1) or 1), 1)

    if columns == 1:
        data = prepare_raw_label(original_data, quantidade)
        send_raw_to_printer(printer_name, data, f"{label_path.name} x{quantidade}")
        record_printed_quantity(label_path.stem, quantidade)
        return

    pair_rows = quantidade // 2
    single_rows = quantidade % 2

    if pair_rows:
        data = prepare_two_column_label(original_data, pair_rows, config)
        send_raw_to_printer(printer_name, data, f"{label_path.name} x{pair_rows * 2}")
        record_printed_quantity(label_path.stem, pair_rows * 2)

    if single_rows:
        data = prepare_raw_label(original_data, single_rows)
        send_raw_to_printer(printer_name, data, f"{label_path.name} x{single_rows}")
        record_printed_quantity(label_path.stem, single_rows)


def prepare_two_column_label(data, rows, config):
    label_match = _last_zpl_label_match(data)
    if not label_match:
        return prepare_raw_label(data, rows * 2)

    label = label_match.group(0)
    pq_match = _last_pq_match(label)
    if not pq_match:
        return prepare_raw_label(data, rows * 2)

    draw_start = _find_draw_start(label[: pq_match.start()])
    if draw_start is None:
        return prepare_raw_label(data, rows * 2)

    prefix = label[:draw_start]
    left_content = label[draw_start: pq_match.start()]
    right_offset_x = max(int(config.get("label_right_x_dots", config.get("two_column_offset_dots", 0))) - int(config.get("label_left_x_dots", 0)), 0)
    right_offset_y = int(config.get("label_right_y_dots", 0)) - int(config.get("label_left_y_dots", 0))
    right_content = _shift_zpl_position(left_content, right_offset_x, right_offset_y)
    label_with_two_columns = prefix + left_content + right_content + label[pq_match.start() :]
    label_with_two_columns = prepare_raw_label(label_with_two_columns, rows)

    return data[: label_match.start()] + label_with_two_columns + data[label_match.end() :]


def prepare_raw_label(data, quantidade):
    if b"^PQ" in data:
        match = _last_pq_match(data)
        if match:
            replacement = f"^PQ{quantidade},0,1,Y".encode("ascii")
            return data[: match.start()] + replacement + data[match.end() :]

    if b"^XZ" in data:
        last_xz = data.rfind(b"^XZ")
        return data[:last_xz] + f"^PQ{quantidade},0,1,Y\r\n".encode("ascii") + data[last_xz:]

    return data


def _last_pq_match(data):
    matches = list(re.finditer(rb"\^PQ\d+(,\d+,\d+,[YN])?", data))
    return matches[-1] if matches else None


def _last_zpl_label_match(data):
    matches = list(re.finditer(rb"\^XA.*?\^XZ", data, flags=re.DOTALL))
    return matches[-1] if matches else None


def _shift_zpl_position(data, x_offset=0, y_offset=0):
    data = re.sub(
        rb"\^FO(\d+),(\d+)",
        lambda match: f"^FO{int(match.group(1)) + x_offset},{int(match.group(2)) + y_offset}".encode("ascii"),
        data,
    )
    data = re.sub(
        rb"\^FT(\d+),(\d+)",
        lambda match: f"^FT{int(match.group(1)) + x_offset},{int(match.group(2)) + y_offset}".encode("ascii"),
        data,
    )
    return data


def _find_draw_start(data):
    positions = [position for position in (data.find(b"^FO"), data.find(b"^FT"), data.find(b"^BY")) if position >= 0]
    return min(positions) if positions else None


def get_default_printer_name():
    winspool = ctypes.WinDLL("winspool.drv")
    needed = ctypes.wintypes.DWORD(0)
    winspool.GetDefaultPrinterW(None, ctypes.byref(needed))

    if needed.value == 0:
        raise RuntimeError("Nenhuma impressora padrao encontrada.")

    buffer = ctypes.create_unicode_buffer(needed.value)
    if not winspool.GetDefaultPrinterW(buffer, ctypes.byref(needed)):
        raise ctypes.WinError()

    return buffer.value


def send_raw_to_printer(printer_name, data, job_name):
    winspool = ctypes.WinDLL("winspool.drv")
    printer_handle = ctypes.wintypes.HANDLE()

    if not winspool.OpenPrinterW(printer_name, ctypes.byref(printer_handle), None):
        raise ctypes.WinError()

    class DOC_INFO_1(ctypes.Structure):
        _fields_ = [
            ("pDocName", ctypes.wintypes.LPWSTR),
            ("pOutputFile", ctypes.wintypes.LPWSTR),
            ("pDatatype", ctypes.wintypes.LPWSTR),
        ]

    doc_info = DOC_INFO_1(job_name, None, "RAW")
    written = ctypes.wintypes.DWORD(0)

    try:
        if not winspool.StartDocPrinterW(printer_handle, 1, ctypes.byref(doc_info)):
            raise ctypes.WinError()
        try:
            if not winspool.StartPagePrinter(printer_handle):
                raise ctypes.WinError()
            buffer = ctypes.create_string_buffer(data)
            if not winspool.WritePrinter(printer_handle, buffer, len(data), ctypes.byref(written)):
                raise ctypes.WinError()
            winspool.EndPagePrinter(printer_handle)
        finally:
            winspool.EndDocPrinter(printer_handle)
    finally:
        winspool.ClosePrinter(printer_handle)


def serialize_settings(config, paths):
    layout = get_label_layout(config)
    return {
        "data_root": config.get("data_root", ""),
        "reposicao_csv": config.get("reposicao_csv", ""),
        "labels_dir": config.get("labels_dir", ""),
        "report_dir": config.get("report_dir", ""),
        "base_file": config.get("base_file", ""),
        "printer_name": config.get("printer_name", ""),
        "two_column_offset_dots": int(layout["two_column_offset_dots"]),
        "label_width_dots": int(layout["width_dots"]),
        "label_height_dots": int(layout["height_dots"]),
        "label_page_width_dots": int(layout["page_width_dots"]),
        "active_label_profile": layout["id"],
        "label_columns": int(layout["columns"]),
        "source_section_prefix": config.get("source_section_prefix", DEFAULT_CONFIG["source_section_prefix"]),
        "resolved": {
            "data_root": str(paths["data_root"]),
            "reposicao_csv": str(paths["reposicao_csv"]),
            "labels_dir": str(paths["labels_dir"]),
            "report_dir": str(paths["report_dir"]),
            "base_file": str(paths["base_file"]),
        },
        "exists": {
            "reposicao_csv": paths["reposicao_csv"].is_file(),
            "labels_dir": paths["labels_dir"].is_dir(),
            "report_dir": paths["report_dir"].is_dir(),
            "base_file": paths["base_file"].is_file(),
        },
    }


def serialize_label_config(config):
    layout = get_label_layout(config)
    profiles = get_label_profiles(config)
    return {
        "dpi": LABEL_DPI,
        "active_profile": layout["id"],
        "profiles": {
            profile_id: {
                "id": profile_id,
                "name": profile["name"],
                "width_mm": profile["width_mm"],
                "height_mm": profile["height_mm"],
                "columns": profile["columns"],
                "gap_mm": profile["gap_mm"],
                "left_x_mm": profile["left_x_mm"],
                "left_y_mm": profile["left_y_mm"],
                "right_x_mm": profile["right_x_mm"],
                "right_y_mm": profile["right_y_mm"],
                "printer_name": profile["printer_name"],
                "width_dots": mm_to_dots(profile["width_mm"]),
                "height_dots": mm_to_dots(profile["height_mm"]),
                "left_x_dots": mm_to_dots(profile["left_x_mm"]),
                "left_y_dots": mm_to_dots(profile["left_y_mm"]),
                "right_x_dots": mm_to_dots(profile["right_x_mm"]),
                "right_y_dots": mm_to_dots(profile["right_y_mm"]),
                "two_column_offset_dots": max(mm_to_dots(profile["right_x_mm"]) - mm_to_dots(profile["left_x_mm"]), 0),
            }
            for profile_id, profile in profiles.items()
        },
        "active_profile_data": {
            "id": layout["id"],
            "name": layout["name"],
            "width_mm": layout["width_mm"],
            "height_mm": layout["height_mm"],
            "columns": layout["columns"],
            "gap_mm": layout["gap_mm"],
            "left_x_mm": layout["left_x_mm"],
            "left_y_mm": layout["left_y_mm"],
            "right_x_mm": layout["right_x_mm"],
            "right_y_mm": layout["right_y_mm"],
            "printer_name": layout["printer_name"],
            "width_dots": layout["width_dots"],
            "height_dots": layout["height_dots"],
            "left_x_dots": layout["left_x_dots"],
            "left_y_dots": layout["left_y_dots"],
            "right_x_dots": layout["right_x_dots"],
            "right_y_dots": layout["right_y_dots"],
            "gap_dots": layout["gap_dots"],
            "page_width_dots": layout["page_width_dots"],
            "two_column_offset_dots": layout["two_column_offset_dots"],
            "column_positions": layout["column_positions"],
        },
    }


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse(request, "login.html")

@app.get("/tela_inicial", response_class=HTMLResponse)
def tela_inicial(request: Request):
    return templates.TemplateResponse(request, "tela_inicial.html")

@app.get("/cartao-trello", response_class=HTMLResponse)
def cartao_trello(request: Request):
    return templates.TemplateResponse(request, "trello.html")

@app.get("/api/produto")
def buscar_produto(codigo: str):
    from trello import produto as descricao_do_Produto

    descricao = descricao_do_Produto(codigo)

    if descricao is None:
        return {"erro": "Produto não encontrado"}

    return {
        "descricao": str(descricao["descricao"]),
        "opcoes": descricao['opcoes']
    }

@app.post("/api/enviar-para-trello")
def enviar_cartao_trello(dados: DadosCartao):

    from trello import executar
    
    enviar_trello = executar(dados.codigo, dados.op, dados.quantidade, dados.linhaCelula)

    if enviar_trello:
        return {"messagem": "Cartão enviado para o Trello com sucesso."}

    return {'erro': 'Não foi possível enviar o cartão parao Trello'}

    
    


@app.get("/reposicao", response_class=HTMLResponse)
def reposicao_page(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.get("/api/itens")
def api_items():
    config = get_config()
    reposicao = get_reposicao_metadata(config)
    items = get_items()
    total_vendido = sum(numeric_quantity(item.get("quantidade_vendida")) for item in items)
    return {
        "items": items,
        "total_items": len(items),
        "total_vendido": total_vendido,
        "reposicao": reposicao,
        "access_url": get_access_url(),
    }


@app.get("/api/base/{codigo}")
def api_base_item(codigo: str):
    item = get_base_item(codigo)
    if item is None:
        raise HTTPException(status_code=404, detail="Codigo nao encontrado na base.")
    return {"item": item}


@app.get("/api/printers")
def api_printers():
    try:
        printers = list_available_printers()
    except Exception as error:
        printers = []
        default_printer = ""
        try:
            default_printer = get_default_printer_name()
        except Exception:
            default_printer = ""
        return {
            "printers": [default_printer] if default_printer else [],
            "default_printer": default_printer,
            "error": str(error),
        }

    default_printer = ""
    try:
        default_printer = get_default_printer_name()
    except Exception:
        default_printer = printers[0] if printers else ""

    return {"printers": printers, "default_printer": default_printer}


@app.get("/config-etiquetas", response_class=HTMLResponse)
def label_config_page(request: Request):
    return templates.TemplateResponse(request, "label_config.html")


@app.get("/api/label-config")
def api_get_label_config():
    config = get_config()
    return {"config": serialize_label_config(config), "access_url": get_access_url()}


@app.post("/api/label-config")
def api_set_label_config(payload: LabelConfigRequest):
    current = get_config()
    profiles = get_label_profiles(current)

    for profile_id, profile_payload in payload.profiles.items():
        if profile_id not in profiles:
            continue

        profiles[profile_id] = normalize_label_profile(profile_id, profile_payload.model_dump())

    active_profile = payload.active_profile.strip()
    if active_profile not in profiles:
        raise HTTPException(status_code=400, detail="Perfil de etiqueta invalido.")

    updated = {
        **current,
        "active_label_profile": active_profile,
        "label_profiles": {
            profile_id: {key: profile[key] for key in LABEL_STORAGE_KEYS}
            for profile_id, profile in profiles.items()
        },
    }

    try:
        set_config(updated)
        return {
            "message": "Configuracao de etiquetas salva com sucesso.",
            "config": serialize_label_config(get_config()),
        }
    except (OSError, ValueError) as error:
        raise HTTPException(status_code=400, detail=str(error))


@app.post("/api/label-config/reset")
def api_reset_label_config():
    updated = {**DEFAULT_CONFIG}
    set_config(updated)
    return {
        "message": "Configuracao de etiquetas restaurada.",
        "config": serialize_label_config(get_config()),
    }


@app.get("/api/config")
def api_get_config():
    config = get_config()
    paths = get_paths(config)
    return {"config": serialize_settings(config, paths), "reposicao": get_reposicao_metadata(config), "access_url": get_access_url()}


@app.post("/api/config")
def api_set_config(payload: ConfigRequest):
    current = get_config()
    data = payload.model_dump(exclude_none=True)
    updated = {**current}
    updated.update(data)
    try:
        paths = get_paths(updated)
        ensure_parent_dirs(paths)
        set_config(updated)
        return {"message": "Configuracao salva com sucesso.", "config": serialize_settings(get_config(), get_paths())}
    except (OSError, ValueError) as error:
        raise HTTPException(status_code=400, detail=str(error))


@app.post("/api/config/reset")
def api_reset_config():
    set_config(DEFAULT_CONFIG)
    paths = get_paths()
    ensure_parent_dirs(paths)
    return {"message": "Configuracao restaurada.", "config": serialize_settings(get_config(), paths)}


@app.post("/api/imprimir")
def api_print(payload: PrintRequest, background_tasks: BackgroundTasks):
    item = next((item for item in get_items() if item["codigo"] == payload.etiqueta), None)
    if item is None:
        item = get_base_item(payload.etiqueta)
    if item is None:
        raise HTTPException(status_code=404, detail="Codigo nao encontrado no csv Reposicao e Diversos nem na base.")
    return queue_print_item(background_tasks, item, payload.quantidade)


@app.post("/api/imprimir-todos")
def api_print_all(background_tasks: BackgroundTasks):
    sold_items = [item for item in get_items() if numeric_quantity(item.get("quantidade_vendida")) > 0 and item.get("automatico")]
    if not sold_items:
        return {"message": "Nenhum item vendido disponivel para impressao.", "total_enviados": 0, "total_itens": 0, "itens": []}

    background_tasks.add_task(print_sold_items_batch)
    return {
        "message": "Impressao de todos os itens vendidos enfileirada.",
        "total_enviados": len(sold_items),
        "total_itens": len(sold_items),
        "itens": [
            {
                "codigo": item["codigo"], "quantidade": numeric_quantity(item.get("quantidade_vendida"))
                } 
                for item in sold_items
                ],
    }




if __name__ == "__main__":
    import uvicorn

    SERVER_PORT = find_available_port(8000, 8002)
    os.environ["PORT"] = str(SERVER_PORT)
    print(f"Servidor disponivel em http://{get_lan_ip()}:{SERVER_PORT}")
    uvicorn.run(app, host="0.0.0.0", port=SERVER_PORT, reload=False)
