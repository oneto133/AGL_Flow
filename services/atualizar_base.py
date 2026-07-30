import pandas as pd                 
import openpyxl as xl

caminho_base = r"relatorio/base.xlsx"

async def atualizar_base_reposicao(codigo, descricao, ean):
    """Função para atualizar a base de dados de reposicao"""
    try:
        wb = xl.load_workbook(caminho_base)

        sheet = wb.active

        proxima_linha = sheet.max_row + 1

        item = [codigo, descricao, ean]

        for c, d in enumerate(item, start=1):
            sheet.cell(row=proxima_linha, column = c, value = d)

        wb.save(caminho_base)

        return {'status': 'sucesso', 'mensagem': f'Produto salvo com sucesso na base'}

    except PermissionError as e:
        return {'status': 'erro', 'mensagem': f'Não foi possível salvar o produto na base'}
