from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from copy import copy
from tempfile import NamedTemporaryFile

from config import CSV_DIR
from utils import _ler_csv, adicionar_horas_uteis, calcular_horas_uteis, normalize_text


SEQUENCIAMENTO = CSV_DIR / "sequenciamento.csv"
APONTAMENTO = CSV_DIR / "apontamento.csv"
CONFIG_LINHAS = CSV_DIR / "config_linhas.csv"
CRONOANALISE = CSV_DIR / "cronoanalise.csv"
BASCULANTES_XLSX = CSV_DIR.parent / "xlsx" / "basculantes.xlsx"
BASCULANTES_CSV = CSV_DIR / "basculantes.csv"
MODELO_BASCULANTE_XLSX = CSV_DIR.parent / "xlsx" / "modelo_impressao_basculante.xlsx"

FAIXAS_MINUTOS = {
    "deslizante": (3.0, 7.0),
    "pivotante": (7.0, 20.0),
    "basculante": (5.0, 10.0),
    "newbv": (5.0, 10.0),
}
ROTULOS = {"deslizante": "Deslizante", "pivotante": "Pivotante", "basculante": "Basculante", "newbv": "NEW BV"}


def _texto(valor: Any) -> str:
    return "" if pd.isna(valor) else str(valor).strip()


def _numero(valor: Any) -> float:
    if valor is None:
        return 0.0
    try:
        if pd.isna(valor):
            return 0.0
    except (TypeError, ValueError):
        return 0.0
    try:
        return float(str(valor).strip().replace(",", "."))
    except (TypeError, ValueError):
        convertido = pd.to_numeric(valor, errors="coerce")
        return 0.0 if pd.isna(convertido) else float(convertido)


def _data(valor: Any) -> pd.Timestamp:
    texto = _texto(valor)
    if not texto:
        return pd.NaT
    try:
        if len(texto) >= 4 and texto[:4].isdigit() and texto[4:5] == "-":
            return pd.Timestamp(texto)
        for formato in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return pd.Timestamp(datetime.strptime(texto, formato))
            except ValueError:
                pass
    except (TypeError, ValueError):
        pass
    return pd.to_datetime(texto, errors="coerce", dayfirst=True)


def _data_exibicao(valor: Any) -> str:
    data = _data(valor)
    if pd.isna(data):
        return "-"
    return data.strftime("%d/%m/%Y %H:%M:%S")


def dados_impressao_basculante(op: str, codigo: str, linha: str = "") -> dict[str, Any] | None:
    seq, _, config = _carga()
    config_map = {normalize_text(row["celula_linha"]): row for _, row in config.iterrows()}
    linha_chave = normalize_text(linha)
    candidatos = seq.loc[seq["op"].astype(str).str.strip() == str(op).strip()]
    if linha_chave:
        candidatos = candidatos.loc[candidatos["linha"].astype(str).map(normalize_text) == linha_chave]
    if candidatos.empty:
        return None

    ordem = candidatos.iloc[-1]
    conf = config_map.get(normalize_text(ordem.get("linha", linha)), {})
    secao = normalize_text(conf.get("secao", "")) if hasattr(conf, "get") else ""
    if secao != "basculante":
        return None

    codigo_texto = str(codigo or ordem.get("codigo_produto", "")).strip()
    produto = {}
    if BASCULANTES_CSV.exists():
        base_basculantes = pd.read_csv(BASCULANTES_CSV, encoding="utf-8-sig", dtype=str).fillna("")
        base_basculantes.columns = [str(col).strip() for col in base_basculantes.columns]
        registro = base_basculantes.loc[
            base_basculantes["CÓD DO PRODUTO"].astype(str).str.strip() == codigo_texto
        ]
        if not registro.empty:
            valores = registro.iloc[0]
            produto = {
                "descricao": valores.get("DESCRIÇÃO", ""),
                "central": valores.get("CENTRAL", ""),
                "carenagem": valores.get("CARENAGEM", ""),
                "estator": valores.get("ESTATOR E SEM FIM", ""),
                "capacitor": valores.get("CAPACITOR", ""),
            }

    return {
        "titulo": "LINHA BASCULANTE ANTIGO",
        "emissao": pd.Timestamp.now().strftime("%d/%m/%Y"),
        "codigo": codigo_texto,
        "quantidade": int(_numero(ordem.get("quantidade"))),
        "descricao": _texto(produto.get("descricao")) or _texto(ordem.get("descricao_produto")),
        "central": _texto(produto.get("central")),
        "carenagem": _texto(produto.get("carenagem")),
        "estator": _texto(produto.get("estator")),
        "capacitor": _texto(produto.get("capacitor")),
        "op": _texto(ordem.get("op")),
        "data_etiq": "--",
        "aj_est": "",
        "prev_entrega": _data_exibicao(ordem.get("previsao_entrada")),
    }


def gerar_planilha_impressao_basculante(op: str, codigo: str, linha: str = "") -> str | None:
    """Gera uma cópia temporária do modelo Excel sem alterar o original."""
    dados = dados_impressao_basculante(op, codigo, linha)
    if not dados or not MODELO_BASCULANTE_XLSX.exists():
        return None

    arquivo_temporario = NamedTemporaryFile(prefix="impressao_basculante_", suffix=".xlsx", delete=False)
    caminho = arquivo_temporario.name
    arquivo_temporario.close()
    try:
        workbook = load_workbook(MODELO_BASCULANTE_XLSX)
        sheet = workbook.active
        for row in range(4, 24):
            for column in range(1, 12):
                sheet.cell(row=row, column=column).value = None
        sheet.delete_rows(18, 3)

        sheet["C3"] = f"EMISSÃO: {dados['emissao']}"
        valores = [
            dados["codigo"], dados["quantidade"], dados["descricao"], dados["central"],
            dados["carenagem"], dados["estator"], dados["capacitor"], dados["op"],
            dados["data_etiq"], dados["aj_est"], dados["prev_entrega"],
        ]
        for column, valor in enumerate(valores, start=1):
            cell = sheet.cell(row=6, column=column)
            cell.value = valor
            alinhamento = copy(cell.alignment)
            alinhamento.shrink_to_fit = True
            cell.alignment = alinhamento
            if column == 11:
                fonte = copy(cell.font)
                fonte.bold = True
                cell.font = fonte

        sheet.column_dimensions["K"].width = 20
        sheet.column_dimensions["C"].width = 58

        sheet["B18"] = "=SUM(B5:B17)"
        sheet.print_area = "A1:K18"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.print_options.horizontalCentered = True
        workbook.save(caminho)
        workbook.close()
        return caminho
    except Exception:
        try:
            Path(caminho).unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _carga() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    seq = _ler_csv(SEQUENCIAMENTO)
    apont = _ler_csv(APONTAMENTO)
    config = _ler_csv(CONFIG_LINHAS)
    for df, cols in ((seq, ["op", "codigo_produto", "descricao_produto", "quantidade", "linha", "operador", "status", "fila", "data_hora_sequenciamento", "data_hora_finalizacao", "hora_inicio_fila", "previsao_entrada", "tempo_total_fila"]), (apont, ["op", "codigo", "quantidade", "data_hora", "status", "observacao"]), (config, ["celula_linha", "secao"])):
        for col in cols:
            if col not in df.columns:
                df[col] = ""
    return seq, apont, config


def _historico(seq: pd.DataFrame, linha: str | None = None, secao: str | None = None, codigo: str | None = None) -> pd.Series:
    df = seq.attrs.get("_historico_cache")
    if df is None:
        df = seq.copy()
        df["qtd"] = pd.to_numeric(df["quantidade"], errors="coerce").fillna(0)
        df["tempo"] = pd.to_numeric(df["tempo_total_fila"], errors="coerce").fillna(0) * 60
        df["status_normalizado"] = df["status"].astype(str).map(normalize_text)
        df["linha_normalizada"] = df["linha"].astype(str).map(normalize_text)
        df["secao_normalizada"] = df["secao"].astype(str).map(normalize_text)
        df["codigo_chave"] = df["codigo_produto"].astype(str).str.strip()
        seq.attrs["_historico_cache"] = df
    filtro = (df["status_normalizado"] == "concluido") & (df["qtd"] > 0) & (df["tempo"] > 0)
    if linha:
        filtro &= df["linha_normalizada"] == normalize_text(linha)
    if secao:
        filtro &= df["secao_normalizada"] == normalize_text(secao)
    if codigo:
        filtro &= df["codigo_chave"] == str(codigo).strip()
    return (df.loc[filtro, "tempo"] / df.loc[filtro, "qtd"]).dropna()


def _ritmo_historico(seq: pd.DataFrame, linha: str, secao: str, codigo: str) -> dict[str, Any] | None:
    """Obtém minutos por unidade sem criar uma previsão com poucos dados."""
    capacidade = _capacidade(seq, linha, secao, codigo)
    if not capacidade.get("amostras"):
        return None
    return {
        "minutos_por_unidade": capacidade["minutos_por_unidade"],
        "amostras": capacidade["amostras"],
        "origem": capacidade["origem_confianca"],
    }


def calcular_metricas_producao(
    seq: pd.DataFrame,
    row: pd.Series,
    apontamentos: pd.DataFrame,
    linha: str,
    secao: str,
    agora: pd.Timestamp | None = None,
    capacidade: dict[str, Any] | None = None,
    crono: pd.DataFrame | None = None,
) -> dict[str, Any]:
    """Calcula quanto deveria ter sido produzido até o momento atual."""
    inicio = _data(row.get("hora_inicio_fila"))
    ritmo_crono = _taxa_cronoanalise(
        crono if crono is not None else pd.DataFrame(),
        linha,
        secao,
        _texto(row.get("codigo_produto")),
        _texto(row.get("descricao_produto")),
    )
    ritmo = (
        {"minutos_por_unidade": 60 / ritmo_crono[0], "amostras": 0, "origem": ritmo_crono[1]}
        if ritmo_crono and ritmo_crono[0] > 0
        else capacidade or _ritmo_historico(
            seq,
            linha=linha,
            secao=secao,
            codigo=_texto(row.get("codigo_produto")),
        )
    )

    fim = agora or pd.Timestamp.now()
    inicio_dia = fim.normalize() + pd.Timedelta(hours=7)
    inicio_producao = min(fim, inicio_dia)
    registros_com_data = apontamentos.copy()
    if "data" not in registros_com_data.columns:
        registros_com_data["data"] = registros_com_data["data_hora"].map(_data)
    registros_dia = registros_com_data.loc[
        registros_com_data["data"].notna()
        & (registros_com_data["data"] >= inicio_dia)
        & (registros_com_data["data"] <= fim)
    ]
    quantidade_apontada_dia = int(registros_dia["quantidade"].map(_numero).sum()) if not registros_dia.empty else 0

    base = {
        "disponivel": False,
        "mensagem": "Dados insuficientes para gerar métricas.",
        "quantidade_deveria_produzida": 0,
        "quantidade_apontada": int(apontamentos["quantidade"].map(_numero).sum()) if not apontamentos.empty else 0,
        "quantidade_apontada_dia": quantidade_apontada_dia,
        "desempenho_percentual": 0,
        "diferenca": None,
        "minutos_por_unidade": None,
        "horas_uteis_decorridas": None,
        "amostras": ritmo["amostras"] if ritmo else 0,
        "origem_ritmo": ritmo.get("origem", ritmo.get("origem_confianca")) if ritmo else None,
    }

    if pd.isna(inicio) or ritmo is None or ritmo["minutos_por_unidade"] <= 0:
        return base

    minutos_decorridos = min(9 * 60, max(0, calcular_horas_uteis(inicio_producao, fim) * 60))
    programada = _numero(row.get("quantidade"))
    deveria_calculado = int((minutos_decorridos / ritmo["minutos_por_unidade"]) if ritmo["minutos_por_unidade"] > 0 else 0)
    deveria = min(programada, deveria_calculado) if programada > 0 else 0
    apontada = base["quantidade_apontada_dia"]

    base.update(
        {
            "disponivel": True,
            "mensagem": "Métrica calculada com base no histórico.",
            "quantidade_deveria_produzida": max(0, deveria),
            "desempenho_percentual": round((apontada / deveria) * 100, 1) if deveria > 0 else 0,
            "diferenca": apontada - max(0, deveria),
            "minutos_por_unidade": ritmo["minutos_por_unidade"],
            "horas_uteis_decorridas": minutos_decorridos / 60,
            "origem_ritmo": ritmo.get("origem", ritmo.get("origem_confianca")),
        }
    )
    return base


def _previsao_saida(
    seq: pd.DataFrame,
    row: pd.Series,
    capacidade: dict[str, Any],
    crono: pd.DataFrame | None = None,
) -> pd.Timestamp:
    """Usa a entrada da próxima OP ou calcula respeitando o expediente útil."""
    inicio = _data(row.get("hora_inicio_fila"))
    if pd.isna(inicio):
        return pd.NaT

    linha_atual = _texto(row.get("linha"))
    fila_atual = _numero(row.get("fila"))
    candidatos = seq.loc[
        (seq["linha"].astype(str).str.strip() == linha_atual)
        & (seq["status"].map(normalize_text) != "concluido")
        & seq["fila"].notna()
    ].copy()
    candidatos["fila_num"] = candidatos["fila"].map(_numero)
    proximos = candidatos.loc[candidatos["fila_num"] > fila_atual].sort_values("fila_num")

    if not proximos.empty:
        entrada_proxima = _data(proximos.iloc[0].get("hora_inicio_fila"))
        if not pd.isna(entrada_proxima) and entrada_proxima > inicio:
            return entrada_proxima

    taxa_crono = _taxa_cronoanalise(
        crono if crono is not None else pd.DataFrame(),
        linha_atual,
        _texto(row.get("secao", "")),
        _texto(row.get("codigo_produto", "")),
        _texto(row.get("descricao_produto", "")),
    )
    minutos_por_unidade = 60 / taxa_crono[0] if taxa_crono and taxa_crono[0] > 0 else capacidade.get("minutos_por_unidade", 0)
    if minutos_por_unidade <= 0 or (not taxa_crono and not capacidade.get("amostras")):
        return pd.NaT

    apontada = _numero(row.get("quantidade_produzida", 0))
    quantidade = _numero(row.get("quantidade"))
    restante = max(quantidade - apontada, 0)
    horas_restantes = restante * minutos_por_unidade / 60
    return adicionar_horas_uteis(inicio, horas_restantes)


def _capacidade(seq: pd.DataFrame, linha: str, secao: str, codigo: str) -> dict[str, Any]:
    faixa = FAIXAS_MINUTOS.get(normalize_text(secao).replace(" ", ""))
    for nome, filtro in (("produto", {"codigo": codigo}), ("secao", {"secao": secao}), ("linha", {"linha": linha})):
        valores = _historico(seq, **filtro)
        if faixa:
            valores = valores.loc[(valores >= faixa[0]) & (valores <= faixa[1])]
        if len(valores) >= 3:
            nivel = "alta" if nome == "produto" else "média" if nome == "secao" else "baixa"
            return {"minutos_por_unidade": round(float(valores.mean()), 2), "confianca": nivel, "origem_confianca": nome, "amostras": int(len(valores))}
    minimo, maximo = faixa or (5.0, 10.0)
    return {"minutos_por_unidade": (minimo + maximo) / 2, "confianca": "baixa", "origem_confianca": "faixa padrão", "amostras": 0}


def _apontamentos_da_op(apont: pd.DataFrame, op: str) -> pd.DataFrame:
    df = apont.loc[apont["op"].astype(str).str.strip() == str(op).strip()].copy()
    df["data"] = df["data_hora"].map(_data)
    return df.sort_values("data", na_position="last")


def _apontamentos_por_hora(apontamentos: pd.DataFrame, periodo: str = "dia") -> list[dict[str, Any]]:
    if apontamentos.empty:
        return []

    df = apontamentos.copy()
    df["data"] = df["data_hora"].map(_data)
    df["quantidade_num"] = df["quantidade"].map(_numero)
    if "observacao" not in df.columns:
        df["observacao"] = ""
    df = df.loc[df["data"].notna()].copy()
    if df.empty:
        return []

    df["hora"] = df["data"].dt.floor("h")
    agrupado = df.groupby("hora", as_index=False).agg(
        quantidade=("quantidade_num", "sum"),
        observacao=("observacao", lambda valores: " | ".join(dict.fromkeys(
            str(valor).strip() for valor in valores if str(valor).strip()
        )),
    )).sort_values("hora")
    return [
        {
            "hora": row["hora"].strftime("%d/%m %H:00"),
            "hora_iso": row["hora"].isoformat(),
            "quantidade": int(row["quantidade"]),
            "observacao": str(row["observacao"] or "").strip(),
        }
        for _, row in agrupado.iterrows()
    ]


def _preparar_cronoanalise(crono: pd.DataFrame) -> pd.DataFrame:
    if crono.empty:
        return pd.DataFrame(columns=["secao_normalizada", "linha_normalizada", "produto_normalizado", "taxa_hora"])

    df = crono.copy()
    for coluna in ("secao", "linha", "produto", "tempo_mensurado", "quantidade", "itens_hora", "itens_dia"):
        if coluna not in df.columns:
            df[coluna] = ""
    df["secao_normalizada"] = df["secao"].astype(str).map(lambda valor: normalize_text(valor).replace(" ", ""))
    df["linha_normalizada"] = df["linha"].astype(str).map(normalize_text)
    df["produto_normalizado"] = df["produto"].astype(str).map(normalize_text)
    df["tempo_num"] = df["tempo_mensurado"].map(_numero)
    df["quantidade_num"] = df["quantidade"].map(_numero)
    df["itens_hora_num"] = df["itens_hora"].map(_numero)
    df["itens_dia_num"] = df["itens_dia"].map(_numero)

    # Usa a medição bruta e os dois indicadores cadastrados. Como itens_dia
    # representa um turno de 9 horas, ele é convertido para itens/hora antes
    # da média.
    taxa_mensurada = (df["quantidade_num"] * 60 / df["tempo_num"].replace(0, pd.NA)).fillna(0)
    taxa_hora = df["itens_hora_num"].where(df["itens_hora_num"] > 0, taxa_mensurada)
    taxa_dia_hora = (df["itens_dia_num"] / 9).where(df["itens_dia_num"] > 0, taxa_mensurada)
    df["taxa_hora"] = pd.concat([taxa_mensurada, taxa_hora, taxa_dia_hora], axis=1).replace(0, pd.NA).mean(axis=1).fillna(0)
    return df.loc[df["taxa_hora"] > 0].copy()


def _taxa_cronoanalise(
    crono: pd.DataFrame,
    linha: str,
    secao: str,
    codigo: str,
    descricao: str,
) -> tuple[float, str] | None:
    if crono.empty:
        return None

    linha_chave = normalize_text(linha)
    secao_chave = normalize_text(secao).replace(" ", "")
    produtos_chave = {normalize_text(codigo), normalize_text(descricao)} - {""}
    filtros = (
        (crono["linha_normalizada"] == linha_chave)
        & crono["produto_normalizado"].isin(produtos_chave)
    )
    if filtros.any():
        return float(crono.loc[filtros, "taxa_hora"].mean()), "cronoanálise da linha e produto"

    filtros = crono["linha_normalizada"] == linha_chave
    if filtros.any():
        return float(crono.loc[filtros, "taxa_hora"].mean()), "cronoanálise da linha"

    filtros = crono["secao_normalizada"] == secao_chave
    if filtros.any():
        return float(crono.loc[filtros, "taxa_hora"].mean()), "cronoanálise da seção"

    return None


def _serializa_op(
    row: pd.Series,
    apont: pd.DataFrame,
    capacidade: dict[str, Any],
    seq: pd.DataFrame,
    linha: str,
    secao: str,
    limite_programacao: pd.Timestamp | None = None,
    crono: pd.DataFrame | None = None,
) -> dict[str, Any]:
    op = _texto(row["op"])
    registros = _apontamentos_da_op(apont, op)
    programada = _numero(row["quantidade"])
    produzida = float(registros["quantidade"].map(_numero).sum()) if not registros.empty else 0
    metricas = calcular_metricas_producao(seq, row, registros, linha, secao, capacidade=capacidade, crono=crono)
    agora = pd.Timestamp.now()
    registros_dia = registros.loc[registros["data"] >= agora.normalize()] if not registros.empty else registros
    quantidade_apontada_dia = metricas["quantidade_apontada_dia"]
    ultima = registros.iloc[-1] if not registros.empty else None
    inicio = _data(row.get("hora_inicio_fila"))
    if limite_programacao is None:
        limite_programacao = pd.Timestamp.now()
    programada_disponivel = pd.isna(inicio) or inicio <= limite_programacao
    quantidade_programada_exibicao = programada if programada_disponivel else 0
    row_com_producao = row.copy()
    row_com_producao["quantidade_produzida"] = produzida
    row_com_producao["secao"] = secao
    previsao = _previsao_saida(seq, row_com_producao, capacidade, crono=crono)
    return {"op": op, "operador": _texto(row.get("operador", "")) or "Não informado", "codigo": _texto(row["codigo_produto"]), "descricao": _texto(row["descricao_produto"]), "quantidade_programada": int(quantidade_programada_exibicao), "quantidade_produzida": int(produzida), "quantidade_apontada": metricas["quantidade_apontada"], "quantidade_apontada_dia": quantidade_apontada_dia, "quantidade_deveria_produzida": metricas["quantidade_deveria_produzida"] if programada_disponivel else 0, "diferenca_eficiencia": metricas["diferenca"] if programada_disponivel else 0, "metricas_disponiveis": metricas["disponivel"] and programada_disponivel, "metricas_mensagem": metricas["mensagem"], "metricas_amostras": metricas["amostras"], "metricas_origem": metricas["origem_ritmo"], "ultima_quantidade": int(_numero(ultima["quantidade"])) if ultima is not None else 0, "ultimo_apontamento": _data_exibicao(ultima["data_hora"]) if ultima is not None else "Nenhum registro", "ultima_observacao": _texto(ultima["observacao"]) if ultima is not None and "observacao" in registros.columns else "", "status": _texto(ultima["status"]) if ultima is not None else (_texto(row["status"]) or "Em fila"), "hora_inicio_fila": _data_exibicao(row.get("hora_inicio_fila")), "inicio_producao": inicio.strftime("%d/%m/%Y %H:%M:%S") if not pd.isna(inicio) else "-", "previsao_saida": previsao.strftime("%d/%m/%Y %H:%M:%S") if not pd.isna(previsao) else "-", "fila": int(_numero(row["fila"])) if _texto(row["fila"]) else None, "data_sequenciamento": _data_exibicao(row["data_hora_sequenciamento"]), "posicao": int(_numero(row["fila"])) if _texto(row["fila"]) else None, "capacidade": capacidade}


def _linhas_ativas() -> list[dict[str, Any]]:
    seq, apont, config = _carga()
    crono = _preparar_cronoanalise(_ler_csv(CRONOANALISE))
    agora = pd.Timestamp.now()
    inicio_dia = agora.normalize()
    datas_apontamentos = apont["data_hora"].map(_data).dropna()
    limite_programacao = agora
    apontamentos_anteriores = datas_apontamentos.loc[datas_apontamentos < inicio_dia]
    if not ((datas_apontamentos >= inicio_dia) & (datas_apontamentos <= agora)).any() and not apontamentos_anteriores.empty:
        limite_programacao = apontamentos_anteriores.max().normalize() + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
    config_map = {normalize_text(row["celula_linha"]): row for _, row in config.iterrows()}
    seq = seq.copy()
    seq["secao"] = seq["linha"].map(
        lambda valor: _texto(config_map.get(normalize_text(valor), {}).get("secao", ""))
        if hasattr(config_map.get(normalize_text(valor), {}), "get") else ""
    )
    ativos = seq.loc[(seq["status"].map(normalize_text) != "concluido") & seq["fila"].notna()].copy()
    ativos["fila_num"] = ativos["fila"].map(_numero)
    ativos = ativos.sort_values(["linha", "fila_num"])
    secoes: dict[str, dict[str, Any]] = {}
    for linha, grupo in ativos.groupby(ativos["linha"].astype(str).str.strip(), sort=False):
        conf = config_map.get(normalize_text(linha), {})
        secao = _texto(conf.get("secao", "outras")) if hasattr(conf, "get") else "outras"
        chave = normalize_text(secao).replace(" ", "") or "outras"
        capacidade = _capacidade(seq, linha, chave, _texto(grupo.iloc[0]["codigo_produto"]))
        item = _serializa_op(grupo.iloc[0], apont, capacidade, seq, linha, chave, limite_programacao, crono)
        data_apontamentos = apont["data_hora"].map(_data)
        apontamentos_linha_dia = apont.loc[
            apont["op"].astype(str).str.strip().isin(seq.loc[seq["linha"].astype(str).str.strip() == linha, "op"].astype(str).str.strip())
            & data_apontamentos.ge(agora.normalize())
            & data_apontamentos.le(agora)
        ]
        item["quantidade_apontada_dia"] = int(apontamentos_linha_dia["quantidade"].map(_numero).sum()) if not apontamentos_linha_dia.empty else 0
        item["linha"] = linha
        item["secao"] = chave
        item["cor"] = _texto(conf.get("cor_cartao", "blue")) if hasattr(conf, "get") else "blue"
        secoes.setdefault(chave, {"secao": chave, "titulo": ROTULOS.get(chave, secao.title()), "linhas": []})["linhas"].append(item)
    for secao in secoes.values():
        observacoes = [
            item for item in secao["linhas"]
            if _texto(item.get("ultima_observacao"))
        ]
        observacoes.sort(
            key=lambda item: _data(item.get("ultimo_apontamento")) or pd.Timestamp.min,
            reverse=True,
        )
        secao["ultima_observacao"] = (
            observacoes[0].get("ultima_observacao", "") if observacoes else ""
        )
    return list(secoes.values())


def painel_resumo() -> dict[str, Any]:
    secoes = _linhas_ativas()
    return {"secoes": secoes, "totais": {"secoes": len(secoes), "linhas": sum(len(s["linhas"]) for s in secoes), "ops": sum(len(s["linhas"]) for s in secoes)}}


def painel_detalhe(linha: str, periodo: str = "dia") -> dict[str, Any]:
    seq, apont, config = _carga()
    crono = _preparar_cronoanalise(_ler_csv(CRONOANALISE))
    config_map = {normalize_text(row["celula_linha"]): row for _, row in config.iterrows()}
    seq = seq.copy()
    seq["secao"] = seq["linha"].map(lambda valor: _texto(config_map.get(normalize_text(valor), {}).get("secao", "")) if hasattr(config_map.get(normalize_text(valor), {}), "get") else "")
    conf = config.loc[config["celula_linha"].astype(str).map(normalize_text) == normalize_text(linha)]
    secao = _texto(conf.iloc[0]["secao"]) if not conf.empty else "outras"
    ops = seq.loc[(seq["linha"].astype(str).str.strip() == linha.strip()) & (seq["status"].map(normalize_text) != "concluido") & seq["fila"].notna()].copy()
    ops["fila_num"] = ops["fila"].map(_numero)
    ops = ops.sort_values("fila_num")
    capacidade = _capacidade(seq, linha, normalize_text(secao).replace(" ", ""), _texto(ops.iloc[0]["codigo_produto"]) if not ops.empty else "")
    ordens = [_serializa_op(row, apont, capacidade, seq, linha, normalize_text(secao).replace(" ", ""), crono=crono) for _, row in ops.iterrows()]
    ops_linha = seq.loc[seq["linha"].astype(str).str.strip() == linha.strip(), "op"].astype(str).str.strip().tolist()
    registros_linha = apont.loc[apont["op"].astype(str).str.strip().isin(ops_linha)].copy()
    agora = pd.Timestamp.now()
    periodo = periodo if periodo in {"dia", "semana"} else "dia"
    inicio_periodo = agora.normalize() if periodo == "dia" else agora.normalize() - pd.Timedelta(days=6)
    registros_grafico = registros_linha.copy()
    registros_grafico["data"] = registros_grafico["data_hora"].map(_data)
    registros_grafico = registros_grafico.loc[registros_grafico["data"].notna() & (registros_grafico["data"] >= inicio_periodo) & (registros_grafico["data"] <= agora)].copy()
    apontamentos_hora = _apontamentos_por_hora(registros_grafico, periodo)
    quantidade_apontada_linha = int(registros_linha["quantidade"].map(_numero).sum()) if not registros_linha.empty else 0
    quantidade_periodo = float(registros_grafico["quantidade"].map(_numero).sum()) if not registros_grafico.empty else 0
    horas_periodo = registros_grafico["data"].dt.floor("h").nunique() if not registros_grafico.empty else 0
    media_por_hora = round(quantidade_periodo / horas_periodo, 1) if horas_periodo else 0
    ordem_atual = ordens[0] if ordens else None
    quantidade_programada_atual = ordem_atual["quantidade_programada"] if ordem_atual else 0
    quantidade_deveria_atual = min(
        quantidade_programada_atual,
        ordem_atual["quantidade_deveria_produzida"],
    ) if ordem_atual and ordem_atual["metricas_disponiveis"] else 0
    quantidade_apontada_atual = int(registros_grafico["quantidade"].map(_numero).sum()) if not registros_grafico.empty else 0
    registros = registros_linha.copy()
    registros["data"] = registros["data_hora"].map(_data)
    registros = registros.sort_values("data", ascending=False, na_position="last").head(20)
    return {"linha": linha, "secao": secao, "capacidade": capacidade, "ordens": ordens, "apontamentos": [{"op": _texto(r["op"]), "codigo": _texto(r["codigo"]), "quantidade": int(_numero(r["quantidade"])), "data_hora": _data_exibicao(r["data_hora"]), "status": _texto(r["status"]), "observacao": _texto(r.get("observacao", ""))} for _, r in registros.iterrows()], "apontamentos_hora": apontamentos_hora, "periodo_grafico": periodo, "resumo_producao": {"quantidade_deveria_produzida": quantidade_deveria_atual, "quantidade_apontada": quantidade_apontada_atual, "quantidade_apontada_linha": quantidade_apontada_linha, "quantidade_programada": quantidade_programada_atual, "media_por_hora": media_por_hora, "metricas_disponiveis": bool(ordem_atual and ordem_atual["metricas_disponiveis"])}}


def _ranking_periodo(
    seq: pd.DataFrame,
    apont: pd.DataFrame,
    config: pd.DataFrame,
    inicio: pd.Timestamp,
    fim: pd.Timestamp,
) -> dict[str, Any]:
    config_map = {normalize_text(row["celula_linha"]): row for _, row in config.iterrows()}
    base_seq = seq[["op", "linha", "operador", "quantidade", "hora_inicio_fila"]].copy()
    base_seq["op_chave"] = base_seq["op"].astype(str).str.strip()
    base_seq["secao"] = base_seq["linha"].map(
        lambda valor: _texto(config_map.get(normalize_text(valor), {}).get("secao", "outras"))
        if hasattr(config_map.get(normalize_text(valor), {}), "get") else "outras"
    )
    base_seq["operador"] = base_seq["operador"].map(_texto).replace("", "Não informado")

    registros = apont.copy()
    registros["data"] = registros["data_hora"].map(_data)
    registros["quantidade_num"] = registros["quantidade"].map(_numero)
    registros = registros.loc[
        registros["data"].notna()
        & (registros["data"] >= inicio)
        & (registros["data"] <= fim)
    ].copy()
    if registros.empty:
        return {"total": 0, "media_por_hora": 0, "linhas": [], "secoes": []}

    registros["op_chave"] = registros["op"].astype(str).str.strip()
    registros = registros.merge(
        base_seq[["op_chave", "linha", "secao", "operador"]],
        on="op_chave",
        how="left",
    )
    registros["linha"] = registros["linha"].fillna("Não informada")
    registros["secao"] = registros["secao"].fillna("outras")
    registros["operador"] = registros["operador"].fillna("Não informado").replace("", "Não informado")
    registros["hora"] = registros["data"].dt.floor("h")

    ops_periodo = registros[["op_chave", "secao"]].drop_duplicates()
    programadas = ops_periodo.merge(
        base_seq[["op_chave", "quantidade", "hora_inicio_fila"]],
        on="op_chave",
        how="left",
    )
    programadas["entrada"] = programadas["hora_inicio_fila"].map(_data)
    programadas = programadas.loc[
        programadas["entrada"].isna() | (programadas["entrada"] <= fim)
    ]
    programadas["programada"] = programadas["quantidade"].map(_numero)
    programadas = programadas.groupby("secao", as_index=False)["programada"].sum()
    resumo_secao = registros.groupby("secao", as_index=False).agg(
        produzido=("quantidade_num", "sum"),
        horas_com_apontamento=("hora", "nunique"),
    )
    resumo_secao = resumo_secao.merge(programadas, on="secao", how="left").fillna({"programada": 0})
    resumo_secao["media_por_hora"] = (
        resumo_secao["produzido"] / resumo_secao["horas_com_apontamento"].clip(lower=1)
    ).round(1)
    resumo_secao["variacao"] = resumo_secao["produzido"] - resumo_secao["programada"]
    resumo_secao = resumo_secao.sort_values("produzido", ascending=False)
    secoes = [
        {
            "secao": _texto(row["secao"]),
            "secao_titulo": ROTULOS.get(normalize_text(row["secao"]).replace(" ", ""), _texto(row["secao"]).title()),
            "programada": int(row["programada"]),
            "produzido": int(row["produzido"]),
            "media_por_hora": float(row["media_por_hora"]),
            "variacao": int(row["variacao"]),
        }
        for _, row in resumo_secao.iterrows()
    ]

    agrupado = registros.groupby(["secao", "linha", "operador"], as_index=False).agg(
        quantidade=("quantidade_num", "sum"),
        horas_com_apontamento=("hora", "nunique"),
    )
    agrupado["media_por_hora"] = (
        agrupado["quantidade"] / agrupado["horas_com_apontamento"].clip(lower=1)
    ).round(1)
    agrupado = agrupado.sort_values(
        ["secao", "quantidade", "linha", "operador"],
        ascending=[True, False, True, True],
    )
    agrupado["rank"] = agrupado.groupby("secao").cumcount() + 1

    linhas = [
        {
            "rank": int(row["rank"]),
            "secao": _texto(row["secao"]),
            "secao_titulo": ROTULOS.get(normalize_text(row["secao"]).replace(" ", ""), _texto(row["secao"]).title()),
            "linha": _texto(row["linha"]),
            "operador": _texto(row["operador"]),
            "quantidade": int(row["quantidade"]),
            "media_por_hora": float(row["media_por_hora"]),
        }
        for _, row in agrupado.iterrows()
    ]
    total = int(registros["quantidade_num"].sum())
    horas = registros["hora"].nunique()
    return {"total": total, "media_por_hora": round(total / max(1, horas), 1), "linhas": linhas, "secoes": secoes}


def painel_exibicao() -> dict[str, Any]:
    seq, apont, config = _carga()
    agora = pd.Timestamp.now()
    inicio_dia = agora.normalize()
    inicio_semana = inicio_dia - pd.Timedelta(days=6)
    datas_apontamentos = apont["data_hora"].map(_data).dropna()
    data_referencia = inicio_dia
    dia_eh_atual = True
    apontamentos_anteriores = datas_apontamentos.loc[datas_apontamentos < inicio_dia]
    if not ((datas_apontamentos >= inicio_dia) & (datas_apontamentos <= agora)).any() and not apontamentos_anteriores.empty:
        data_referencia = apontamentos_anteriores.max().normalize()
        dia_eh_atual = False
    inicio_referencia = data_referencia
    fim_referencia = data_referencia + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
    ranking_dia = _ranking_periodo(seq, apont, config, inicio_referencia, fim_referencia)
    ranking_dia["data_referencia"] = data_referencia.strftime("%d/%m/%Y")
    ranking_dia["dia_eh_atual"] = dia_eh_atual
    return {
        "atualizado_em": agora.strftime("%d/%m/%Y %H:%M:%S"),
        "secoes": painel_resumo()["secoes"],
        "ranking": {
            "dia": ranking_dia,
            "semana": _ranking_periodo(seq, apont, config, inicio_semana, agora),
        },
    }


def eficiencia_historico(secao: str = "", linha: str = "", periodo: str = "total") -> dict[str, Any]:
    seq, apont, config = _carga()
    config_map = {normalize_text(row["celula_linha"]): row for _, row in config.iterrows()}
    seq = seq.copy()
    seq["secao"] = seq["linha"].map(lambda valor: _texto(config_map.get(normalize_text(valor), {}).get("secao", "")) if hasattr(config_map.get(normalize_text(valor), {}), "get") else "")
    seq["data_fim"] = seq["data_hora_finalizacao"].map(_data)
    filtro = seq["status"].map(normalize_text) == "concluido"
    if secao:
        filtro &= seq["secao"].map(normalize_text) == normalize_text(secao)
    if linha:
        filtro &= seq["linha"].astype(str).map(normalize_text) == normalize_text(linha)
    agora = pd.Timestamp.now()
    inicio_periodo = None
    fim_periodo = None
    if periodo == "dia":
        inicio_periodo = agora.normalize()
        fim_periodo = agora
        filtro_dia_atual = filtro & (seq["data_fim"] >= inicio_periodo) & (seq["data_fim"] <= agora)
        if filtro_dia_atual.any():
            filtro = filtro_dia_atual
        else:
            datas_anteriores = seq.loc[
                filtro & (seq["data_fim"] < inicio_periodo), "data_fim"
            ].dropna()
            if not datas_anteriores.empty:
                inicio_periodo = datas_anteriores.max().normalize()
                fim_periodo = inicio_periodo + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
                filtro &= (seq["data_fim"] >= inicio_periodo) & (seq["data_fim"] <= fim_periodo)
            else:
                filtro &= seq["data_fim"].dt.date == agora.date()
    elif periodo == "semana":
        filtro &= seq["data_fim"] >= agora - pd.Timedelta(days=7)
        inicio_periodo = agora - pd.Timedelta(days=7)
        fim_periodo = agora
    elif periodo == "mes":
        filtro &= seq["data_fim"] >= agora - pd.Timedelta(days=30)
        inicio_periodo = agora - pd.Timedelta(days=30)
        fim_periodo = agora
    base = seq.loc[filtro].copy()
    apont["data"] = apont["data_hora"].map(_data)
    apont["qtd"] = apont["quantidade"].map(_numero)
    if periodo == "dia":
        apont = apont.loc[(apont["data"] >= inicio_periodo) & (apont["data"] <= fim_periodo)]
    elif periodo == "semana":
        apont = apont.loc[apont["data"] >= agora - pd.Timedelta(days=7)]
    elif periodo == "mes":
        apont = apont.loc[apont["data"] >= agora - pd.Timedelta(days=30)]
    ops_permitidos = set(base["op"].astype(str).str.strip())
    apont = apont.loc[
        apont["op"].astype(str).str.strip().isin(ops_permitidos)
        & apont["data"].notna()
    ].copy()
    apont["data_dia"] = apont["data"].dt.normalize()
    pontos_data = apont.groupby("data_dia", as_index=False).agg(
        produzido=("qtd", "sum"),
    )
    programado_data = apont[["data_dia", "op"]].copy()
    programado_data["op"] = programado_data["op"].astype(str).str.strip()
    programado_data = programado_data.drop_duplicates()
    programado_data = programado_data.merge(
        base[["op", "quantidade"]].assign(op=lambda frame: frame["op"].astype(str).str.strip()),
        on="op",
        how="left",
    )
    programado_data["quantidade"] = programado_data["quantidade"].map(_numero)
    programado_data = programado_data.groupby("data_dia", as_index=False)["quantidade"].sum()
    pontos_data = pontos_data.merge(programado_data, on="data_dia", how="left").fillna({"quantidade": 0})
    pontos_data["eficiencia"] = pontos_data.apply(
        lambda row: round(row["produzido"] / row["quantidade"] * 100, 1) if row["quantidade"] else 0,
        axis=1,
    )
    pontos_data = [
        {
            "data": row["data_dia"].strftime("%d/%m/%Y"),
            "data_iso": row["data_dia"].isoformat(),
            "programado": int(row["quantidade"]),
            "produzido": int(row["produzido"]),
            "eficiencia": float(row["eficiencia"]),
        }
        for _, row in pontos_data.sort_values("data_dia").iterrows()
    ]
    produzido = apont.groupby(apont["op"].astype(str).str.strip())["qtd"].sum().to_dict()
    linhas = []
    for _, row in base.iterrows():
        op = _texto(row["op"])
        programada = _numero(row["quantidade"])
        feito = float(produzido.get(op, 0))
        linhas.append({"op": op, "secao": _texto(row["secao"]), "linha": _texto(row["linha"]), "operador": _texto(row.get("operador", "")) or "Não informado", "data": row["data_fim"].strftime("%d/%m/%Y") if not pd.isna(row["data_fim"]) else "-", "programado": programada, "produzido": feito, "eficiencia": round((feito / programada * 100), 1) if programada else 0})
    por_linha: dict[str, dict[str, Any]] = {}
    for item in linhas:
        resumo = por_linha.setdefault(item["linha"], {"linha": item["linha"], "programado": 0, "produzido": 0})
        resumo["programado"] += item["programado"]
        resumo["produzido"] += item["produzido"]
    resumo = list(por_linha.values())
    for item in resumo:
        item["eficiencia"] = round(item["produzido"] / item["programado"] * 100, 1) if item["programado"] else 0
    linhas.sort(key=lambda item: item["data"])
    def agrupar(campo: str) -> list[dict[str, Any]]:
        agrupado: dict[str, dict[str, Any]] = {}
        for item in linhas:
            chave = item[campo]
            grupo = agrupado.setdefault(chave, {campo: chave, "programado": 0, "produzido": 0})
            grupo["programado"] += item["programado"]
            grupo["produzido"] += item["produzido"]
        for grupo in agrupado.values():
            grupo["eficiencia"] = round(grupo["produzido"] / grupo["programado"] * 100, 1) if grupo["programado"] else 0
        return list(agrupado.values())
    resumo_secao = agrupar("secao")
    resumo_linha = agrupar("linha")
    resumo_operador = agrupar("operador")
    resumo_exibicao = resumo + [{"linha": f"Seção · {item['secao']}", **item} for item in resumo_secao] + [{"linha": f"Operador · {item['operador']}", **item} for item in resumo_operador]
    return {"filtro": {"secao": secao, "linha": linha, "periodo": periodo}, "pontos": pontos_data, "resumo": resumo_exibicao, "resumo_secao": resumo_secao, "resumo_linha": resumo_linha, "resumo_operador": resumo_operador, "totais": {"programado": sum(item["programado"] for item in linhas), "produzido": sum(item["produzido"] for item in linhas), "eficiencia": round(sum(item["produzido"] for item in linhas) / sum(item["programado"] for item in linhas) * 100, 1) if sum(item["programado"] for item in linhas) else 0}}
